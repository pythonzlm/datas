#encoding=utf-8
from openpyxl import *
from openpyxl.styles import Border,Side,Font
import time
import os
from FormatTime import date_time_chinese

class ParseExcel(object):
    def __init__(self,excel_file_path):
        self.excel_file_path = excel_file_path
        self.workbook = load_workbook(excel_file_path)
        self.font = Font(color=None)
        self.colorDict = {"red":'FFFF3030',"green":'FF008B00'}
        self.sheet = self.workbook.active

    def set_sheet_by_index(self,sheet_index):
        self.sheet = self.get_sheet_by_index(sheet_index)

    def set_sheet_by_name(self,sheet_name):
        self.sheet = self.workbook.get_sheet_by_name(sheet_name)

    def get_default_name(self):
        return self.sheet.title

    def get_sheet_by_name(self,sheet_name):
        self.sheet = self.workbook[sheet_name]
        return self.sheet

    def get_sheet_by_index(self,sheet_index):
        sheet_name = self.workbook.sheetnames[sheet_index]
        self.sheet = self.get_sheet_by_name(sheet_name)
        return self.sheet
    def get_max_row_no(self):
        return self.sheet.max_row

    def get_max_col_no(self):
        return self.sheet.max_column

    def get_min_row_no(self):
        return self.sheet.min_row

    def get_min_col_no(self):
        return self.sheet.min_column

    def get_all_rows(self):
        rows=[]
        for row in self.sheet.iter_rows():
            rows.append(row)
        return rows

    def get_all_cols(self):
        cols=[]
        for col in self.sheet.iter_cols():
            cols.append(col)
        return cols
    def get_single_row(self,row_no):
        return self.sheet.get_all_rows()[row_no]

    def get_single_col(self,col_no):
        return self.sheet.get_all_cols()[col_no]

    def get_cell(self,row_no,col_no):
        return self.sheet.cell(row = row_no,column=col_no)

    def write_cell_content(self,row_no,col_no,content,font=None):
        self.sheet.cell(row = row_no,column = col_no).value = content
        self.workbook.save(self.excel_file_path)

    def write_cell_current_time(self,row_no,col_no):
        self.sheet.cell(row = row_no,column = col_no).value = date_time_chinese()
        self.workbook.save(self.excel_file_path)

    def save_excel_file(self):
        self.workbook.save(self.excel_file_path)



